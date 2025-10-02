# 03_valuta_modello.py
# Questo script esegue una valutazione quantitativa del sistema RAG.
# Itera attraverso un 'golden dataset' predefinito di domande e risposte ideali,
# esegue l'intera pipeline RAG per ogni domanda e calcola la similarità semantica
# tra la risposta generata e quella ideale. Gli elementi in cui il modello risponde
# con "Contesto non sufficiente" vengono registrati nel report ma esclusi dal
# calcolo del punteggio medio finale per una metrica di performance più accurata.

import os
import sys
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from dotenv import load_dotenv
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

# Aggiunge la directory principale del progetto al path di Python per consentire l'importazione di moduli locali.
script_directory = os.path.dirname(os.path.abspath(__file__))
sys.path.append(script_directory)

# Importa i moduli personalizzati del progetto.
from index import product
from writer import writer_ollama, writer_azure_openai
from query_rewriter import rewrite_query

# --- Configurazione del Modello ---
# Questa variabile permette di passare da un LLM all'altro per la valutazione.
# Deve corrispondere alle chiavi usate nel dizionario 'writers' (es. "azure", "ollama").
MODELLO_DA_USARE = "azure"  # Opzioni: "azure", "ollama"

def salva_report_excel(risultati, punteggio_medio_totale, risultati_cat, num_validi, num_totali):
    """
    Salva i risultati della valutazione in un file Excel ben formattato.
    Questa funzione gestisce tutta la formattazione e l'aggregazione dei dati per il report finale.
    """
    # Mantiene l'output della console in italiano.
    print("\nSalvataggio del report su file Excel...")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "RAG Evaluation Report"

        # Definisce le intestazioni per il report Excel, inclusa la nuova colonna "Domanda Riformulata".
        headers = ["Categoria", "Domanda Originale", "Domanda Riformulata", "Risposta Ideale", "Risposta Generata", "Punteggio Similarità"]
        ws.append(headers)
        
        # --- Formattazione Excel ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        # Stile per le righe escluse dal calcolo del punteggio medio.
        excluded_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # --- Popolamento dei Dati ---
        for item in risultati:
            ws.append([
                item["categoria"],
                item["domanda"],
                item["domanda_riformulata"],
                item["risposta_ideale"],
                item["risposta_generata"],
                f'{item["score"]:.4f}'
            ])
            # Evidenzia le righe in cui il modello non ha trovato contesto sufficiente.
            if "contesto non sufficiente" in item["risposta_generata"].lower():
                for cell in ws[ws.max_row]:
                    cell.fill = excluded_fill

        # Applica bordi e a capo automatico a tutte le celle di dati per una migliore leggibilità.
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=2, max_row=len(risultati)+1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border
        
        # Imposta la larghezza delle colonne per un layout migliore.
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 20
        
        # --- Sezione di Riepilogo ---
        # Aggiunge un blocco di riepilogo alla fine del report con le metriche complessive.
        start_row = len(risultati) + 4
        ws.cell(row=start_row, column=1, value="Evaluation Summary").font = Font(bold=True, size=14)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
        
        ws.cell(row=start_row + 1, column=1, value="Model Used:").font = Font(bold=True)
        ws.cell(row=start_row + 1, column=2, value=MODELLO_DA_USARE.upper())
        ws.cell(row=start_row + 2, column=1, value="Evaluated Samples:").font = Font(bold=True)
        ws.cell(row=start_row + 2, column=2, value=f"{num_validi} / {num_totali}")
        ws.cell(row=start_row + 3, column=1, value="Overall Average Score:").font = Font(bold=True)
        ws.cell(row=start_row + 3, column=2, value=f"{punteggio_medio_totale:.4f}").font = Font(bold=True, color="00B050")
        
        ws.cell(row=start_row + 5, column=1, value="Average Scores by Category").font = Font(bold=True, size=12)
        ws.merge_cells(start_row=start_row+5, start_column=1, end_row=start_row+5, end_column=3)
        cat_row = start_row + 6
        for categoria, punteggio in risultati_cat.items():
            ws.cell(row=cat_row, column=1, value=f"Category '{categoria}'")
            ws.cell(row=cat_row, column=2, value=f"{punteggio['media']:.4f} ({punteggio['validi']}/{punteggio['totali']} samples)").font = Font(bold=True)
            cat_row += 1

        # Genera un nome file unico con un timestamp.
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f'report_valutazione_{MODELLO_DA_USARE}_{timestamp}.xlsx'
        wb.save(file_name)
        print(f"✅ Report salvato con successo come '{file_name}'")
        
    except Exception as e:
        print(f"ERRORE durante il salvataggio del report Excel: {e}")


def esegui_rag(domanda: str) -> tuple[str, str]:
    """
    Orchestra l'intera pipeline RAG per una singola domanda.
    Questa versione include il passo di riscrittura della query per migliorare il recupero.

    Args:
        domanda (str): La domanda originale dell'utente dal golden dataset.

    Returns:
        tuple[str, str]: Una tupla contenente la risposta generata e la query riscritta.
    """
    print(f"  - Esecuzione RAG per la domanda: '{domanda}'")
    
    # 1. Riscrittura della Query: Migliora la domanda originale per una ricerca semantica più efficace.
    query_per_ricerca = rewrite_query(domanda, MODELLO_DA_USARE)
    
    # 2. Recupero: Usa la query riscritta per trovare documenti pertinenti dall'indice di ricerca.
    documenti_trovati = product.find_products(context=query_per_ricerca)
    
    # 3. Aumento: Prepara il contesto recuperato per essere iniettato nel prompt finale.
    contesto_per_prompt = "\n\n---\n\n".join([doc.get("content", "") for doc in documenti_trovati]) if documenti_trovati else "Nessun contesto specifico è stato trovato."
    
    # 4. Generazione: Seleziona lo scrittore LLM appropriato e genera la risposta finale.
    writers = {
        "ollama": writer_ollama.write,
        "azure": writer_azure_openai.write # Usa "azure" come chiave
    }
    
    # Il prompt finale usa il contesto recuperato ma la domanda ORIGINALE per garantire
    # che la risposta sia direttamente indirizzata alla richiesta iniziale dell'utente.
    risposta_generata = writers[MODELLO_DA_USARE](productContext=contesto_per_prompt, assignment=domanda)
    
    # Restituisce sia la risposta finale sia la query riscritta per un report dettagliato.
    return risposta_generata, query_per_ricerca


# --- Flusso Principale di Valutazione ---
if __name__ == "__main__":
    print(f"--- INIZIO VALUTAZIONE (MODELLO SELEZIONATO: {MODELLO_DA_USARE.upper()}) ---")
    
    # Carica le variabili d'ambiente (chiavi API, endpoint) dal file .env.
    load_dotenv()
    
    print("Caricamento di golden_dataset.json...")
    try:
        dataset_path = os.path.join(script_directory, 'golden_dataset.json')
        with open(dataset_path, 'r', encoding='utf-8') as f:
            golden_dataset = json.load(f)
        print(f"Dataset caricato con successo con {len(golden_dataset)} esempi.")
    except FileNotFoundError:
        print(f"ERRORE: File '{dataset_path}' non trovato.")
        sys.exit()

    # Carica il modello sentence-transformer usato per calcolare i punteggi di similarità del coseno.
    # Questo modello deve essere consistente per produrre punteggi comparabili.
    print("Caricamento del modello di embedding per il calcolo dei punteggi...")
    model = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
    print("Modello di embedding caricato.")

    # Inizializza le liste per memorizzare i risultati per l'aggregazione finale e l'esportazione.
    risultati_per_export = []
    risultati_valutazione_validi = []
    risultati_per_categoria_validi = {}

    print("\nInizio del ciclo di valutazione sul Golden Dataset...")
    # Itera su ogni coppia domanda-risposta nel golden dataset.
    for i, item in enumerate(golden_dataset):
        categoria = item['categoria']
        domanda_test = item["domanda"]
        risposta_ideale = item["risposta_ideale"]

        print(f"\n--- Valutazione elemento {i+1}/{len(golden_dataset)} (Categoria: {categoria}) ---")
        
        # Inizializza le variabili prima del blocco try per prevenire NameError in caso di eccezioni.
        risposta_generata = "ERRORE SCONOSCIUTO"
        domanda_riformulata = "ERRORE SCONOSCIUTO"

        try:
            # Esegue la pipeline RAG e cattura sia la risposta sia la query riscritta.
            risposta_generata, domanda_riformulata = esegui_rag(domanda_test)
            is_contesto_insufficiente = "contesto non sufficiente" in risposta_generata.lower()

            if is_contesto_insufficiente:
                # Se il modello segnala che il contesto era insufficiente, assegna un punteggio di 0.
                # Questa risposta sarà esclusa dal calcolo del punteggio medio finale.
                print("  - Rilevata risposta 'Contesto non sufficiente'. Esclusa dal calcolo della media.")
                score = 0.0
            else:
                # Se viene generata una risposta valida, calcola la sua similarità semantica con la risposta ideale.
                embedding_generato = model.encode([risposta_generata])
                embedding_ideale = model.encode([risposta_ideale])
                score = cosine_similarity(embedding_generato, embedding_ideale)[0][0]
            
            # Stampa un confronto sulla console per il monitoraggio in tempo reale.
            print("\n" + "="*20 + " CONFRONTO RISPOSTE " + "="*20)
            print(f"DOMANDA           : {domanda_test}")
            print(f"QUERY RISCRITTA   : {domanda_riformulata}") # Aggiunto per visibilità
            print(f"RISPOSTA IDEALE   :\n{risposta_ideale}\n")
            print(f"RISPOSTA GENERATA :\n{risposta_generata.strip()}")
            print("="*62 + "\n")
            print(f"  - Punteggio Similarità Semantica: {score:.4f}")
            
        except Exception as e:
            # Cattura eventuali errori inaspettati durante l'esecuzione della pipeline RAG.
            print(f"\nSi è verificato un errore inaspettato: {e}\n")
            score = 0.0
            # Assicura che le variabili abbiano valori di errore se il blocco try fallisce.
            risposta_generata = f"ERRORE: {e}"
            is_contesto_insufficiente = True

        # Memorizza i risultati dell'elemento corrente per la successiva esportazione.
        risultati_per_export.append({
            "categoria": categoria,
            "domanda": domanda_test,
            "domanda_riformulata": domanda_riformulata,
            "risposta_ideale": risposta_ideale,
            "risposta_generata": risposta_generata.strip(),
            "score": score
        })
        
        # Aggrega i punteggi per le risposte valide (non escluse).
        if not is_contesto_insufficiente:
            risultati_valutazione_validi.append(score)
            if categoria not in risultati_per_categoria_validi:
                risultati_per_categoria_validi[categoria] = []
            risultati_per_categoria_validi[categoria].append(score)

    # --- Calcolo Finale dei Punteggi ---
    # Calcola il punteggio medio complessivo e il punteggio medio per ogni categoria.
    punteggio_medio = sum(risultati_valutazione_validi) / len(risultati_valutazione_validi) if risultati_valutazione_validi else 0
    punteggi_medi_categoria_dettagliati = {}
    for cat in set(item['categoria'] for item in golden_dataset):
        scores_validi = risultati_per_categoria_validi.get(cat, [])
        media = sum(scores_validi) / len(scores_validi) if scores_validi else 0
        totali_in_cat = sum(1 for item in golden_dataset if item['categoria'] == cat)
        punteggi_medi_categoria_dettagliati[cat] = {'media': media, 'validi': len(scores_validi), 'totali': totali_in_cat}

    # --- Stampa il Riepilogo Finale sulla Console ---
    print("\n\n--- RISULTATI FINALI DELLA VALUTAZIONE ---")
    print(f"Modello Valutato: {MODELLO_DA_USARE.upper()}")
    num_totali = len(golden_dataset)
    num_validi = len(risultati_valutazione_validi)
    print(f"Campioni Totali: {num_totali}")
    print(f"Campioni Valutati (escluso 'Contesto non sufficiente'): {num_validi}")
    print(f"Punteggio Medio (sui campioni valutati): {punteggio_medio:.4f}")
    
    print("\n--- ANALISI DETTAGLIATA PER CATEGORIA ---")
    for categoria, dati in punteggi_medi_categoria_dettagliati.items():
        print(f"Categoria '{categoria}': Punteggio Medio = {dati['media']:.4f} ({dati['validi']}/{dati['totali']} campioni valutati)")

    # Salva i risultati dettagliati in un file Excel.
    salva_report_excel(risultati_per_export, punteggio_medio, punteggi_medi_categoria_dettagliati, num_validi, num_totali)