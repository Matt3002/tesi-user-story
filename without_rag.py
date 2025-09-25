import os
import sys
import json
import ollama # Importiamo direttamente ollama
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# --- SOLUZIONE PER L'IMPORT ---
script_directory = os.path.dirname(os.path.abspath(__file__))
sys.path.append(script_directory)

# --- Import dei moduli necessari ---
from dotenv import load_dotenv
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

# --- CONFIGURAZIONE ---
MODELLO_VALUTATO = "OLLAMA (Senza RAG)"
OLLAMA_MODEL_NAME = 'llama3.2:latest'
# ---

# ... (La funzione salva_report_excel rimane invariata) ...
def salva_report_excel(risultati, punteggio_medio_totale, risultati_cat):
    print(f"\n💾 Salvataggio del report Excel per '{MODELLO_VALUTATO}'...")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Report Valutazione SENZA RAG"

        headers = ["Categoria", "Domanda", "Risposta Ideale", "Risposta Generata (Senza RAG)", "Punteggio Similarità"]
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # Colore rosso per distinguerlo
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for item in risultati:
            ws.append([item["categoria"], item["domanda"], item["risposta_ideale"], item["risposta_generata"], f'{item["score"]:.4f}'])

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=2, max_row=len(risultati)+1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 20
        
        start_row = len(risultati) + 4
        ws.cell(row=start_row, column=1, value="Riepilogo Valutazione").font = Font(bold=True, size=14)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
        
        ws.cell(row=start_row + 1, column=1, value="Modello Utilizzato:").font = Font(bold=True)
        ws.cell(row=start_row + 1, column=2, value=MODELLO_VALUTATO)
        
        ws.cell(row=start_row + 2, column=1, value="Punteggio Medio Globale:").font = Font(bold=True)
        ws.cell(row=start_row + 2, column=2, value=f"{punteggio_medio_totale:.4f}").font = Font(bold=True, color="00B050")
        
        ws.cell(row=start_row + 4, column=1, value="Punteggi Medi per Categoria").font = Font(bold=True, size=12)
        ws.merge_cells(start_row=start_row+4, start_column=1, end_row=start_row+4, end_column=2)

        cat_row = start_row + 5
        for categoria, punteggio in risultati_cat.items():
            ws.cell(row=cat_row, column=1, value=f"Categoria '{categoria}'")
            ws.cell(row=cat_row, column=2, value=f"{punteggio:.4f}").font = Font(bold=True)
            cat_row += 1
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f'report_valutazione_SENZA_RAG_{timestamp}.xlsx'
        wb.save(file_name)
        print(f"✅ Report salvato con successo come '{file_name}'")
        
    except Exception as e:
        print(f"❌ ERRORE durante il salvataggio del report Excel: {e}")

# --- MODIFICA CHIAVE QUI ---
def esegui_generazione_diretta(domanda: str) -> str:
    print(f"   - Eseguo la generazione diretta (SENZA RAG, prompt semplice) per: '{domanda}'")
    
    # Usiamo un prompt di sistema generico e semplice
    system_prompt = "Il tuo compito è generare User Story."
    user_prompt = f"Genera una o più user story complete per la seguente richiesta: \"{domanda}\""
    
    try:
        response = ollama.chat(
            model=OLLAMA_MODEL_NAME,
            messages=[
                {'role': 'system', 'content': system_prompt},
                {'role': 'user', 'content': user_prompt},
            ]
        )
        return response['message']['content']
    except Exception as e:
        print(f"Errore durante la chiamata a Ollama: {e}")
        return f"ERRORE: {e}"

# --- FLUSSO DI VALUTAZIONE ---
print(f"--- INIZIO VALUTAZIONE DEL MODELLO {MODELLO_VALUTATO} ---")

load_dotenv()

print("📚 Caricamento del golden_dataset.json...")
try:
    dataset_path = os.path.join(script_directory, 'golden_dataset.json')
    with open(dataset_path, 'r', encoding='utf-8') as f:
        golden_dataset = json.load(f)
    print(f"✅ Dataset caricato con successo! {len(golden_dataset)} esempi trovati.")
except FileNotFoundError:
    print(f"❌ ERRORE: File '{dataset_path}' non trovato.")
    sys.exit()

print("🔎 Caricamento del modello di embedding per la valutazione...")
model = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
print("✅ Modello caricato.")

risultati_per_export = []
risultati_per_categoria = {}

print("\n🚀 Avvio del ciclo di valutazione sul Golden Dataset...")
for i, item in enumerate(golden_dataset):
    categoria = item['categoria']
    domanda_test = item["domanda"]
    risposta_ideale = item["risposta_ideale"]

    print(f"\n--- Valutando l'item {i+1}/{len(golden_dataset)} (Categoria: {categoria}) ---")

    try:
        risposta_generata = esegui_generazione_diretta(domanda_test)
        
        print("\n" + "="*20 + " CONFRONTO RISPOSTE (SENZA RAG) " + "="*20)
        print(f"DOMANDA        : {domanda_test}")
        print(f"RISPOSTA IDEALE  :\n{risposta_ideale}\n")
        print(f"RISPOSTA GENERATA:\n{risposta_generata.strip()}")
        print("="*72 + "\n")
        
        embedding_generato = model.encode([risposta_generata])
        embedding_ideale = model.encode([risposta_ideale])
        score = cosine_similarity(embedding_generato, embedding_ideale)[0][0]
        
        print(f"   - Punteggio di Similarità Semantica: {score:.4f}")
        
    except Exception as e:
        print(f"\n❌ Si è verificato un errore inatteso: {e}\n")
        score = 0.0
        risposta_generata = f"ERRORE: {e}"

    risultati_per_export.append({
        "categoria": categoria,
        "domanda": domanda_test,
        "risposta_ideale": risposta_ideale,
        "risposta_generata": risposta_generata.strip(),
        "score": score
    })
    
    if categoria not in risultati_per_categoria:
        risultati_per_categoria[categoria] = []
    risultati_per_categoria[categoria].append(score)

punteggio_medio = sum(score['score'] for score in risultati_per_export) / len(risultati_per_export) if risultati_per_export else 0
punteggi_medi_categoria = {cat: sum(scores) / len(scores) for cat, scores in risultati_per_categoria.items()}

print("\n\n--- RISULTATI FINALI DELLA VALUTAZIONE ---")
print(f"Modello Valutato: {MODELLO_VALUTATO}")
print(f"📈 Punteggio Medio di Similarità Semantica: {punteggio_medio:.4f}")
print("\n--- ANALISI DETTAGLIATA PER CATEGORIA ---")
for categoria, punteggio in punteggi_medi_categoria.items():
    print(f"📊 Categoria '{categoria}': Punteggio Medio = {punteggio:.4f} ({len(risultati_per_categoria[categoria])} campioni)")

salva_report_excel(risultati_per_export, punteggio_medio, punteggi_medi_categoria)
