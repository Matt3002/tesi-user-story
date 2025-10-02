# 04_valutazione_con_LLM.py
# Versione aggiornata con la logica di formattazione fornita dall'utente.

import os
import sys
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
from openai import AzureOpenAI
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Configurazione del Percorso ---
script_directory = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.dirname(script_directory))

nome_file_report = "report_valutazione_ollama_20250930_124300.xlsx"
FILE_REPORT_DA_VALUTARE = os.path.join(script_directory, nome_file_report)

# Carica le variabili d'ambiente.
load_dotenv()

# Inizializza il client per il modello "giudice".
try:
    judge_client = AzureOpenAI(
        api_key=os.getenv("AZURE_OPENAI_API_KEY"),
        api_version="2024-02-01",
        azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT")
    )
    JUDGE_MODEL_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "gpt-4o")
except Exception as e:
    print(f"Errore nell'inizializzazione del client di Azure: {e}")
    sys.exit(1)

def ottieni_giudizio_dettagliato_llm(domanda: str, risposta_ideale: str, risposta_generata: str) -> dict:
    """
    Interroga un LLM "giudice" per ottenere una valutazione dettagliata su 5 metriche.
    """
    print(f"  - Richiesta di giudizio dettagliato per la domanda: '{domanda[:40]}...'")

    system_prompt = """
    Sei un valutatore esperto di sistemi RAG, specializzato in Ingegneria del Software e metodologie Agili.
    Il tuo compito è fornire una valutazione dettagliata della "Risposta Generata" rispetto a una "Risposta Ideale",
    considerando la "Domanda Originale".

    Valuta la risposta su 5 criteri, assegnando un punteggio intero da 1 (pessimo) a 5 (eccellente) per ciascuno.
    Fornisci la tua valutazione SOLO in formato JSON, con le seguenti chiavi:
    1. "punteggio_chiarezza": La risposta è chiara, specifica e inequivocabile?
    2. "punteggio_rilevanza": La risposta è pertinente all'intento della domanda?
    3. "punteggio_correttezza": La risposta è fattualmente accurata rispetto alla risposta ideale?
    4. "punteggio_completezza": La risposta è completa e affronta tutti gli aspetti (anche impliciti) della domanda?
    5. "punteggio_coerenza": Tono, stile e struttura della risposta sono uniformi e appropriati?
    6. "giustificazione_generale": Una frase concisa (massimo 20 parole) che riassume il tuo giudizio complessivo.

    Se la "Risposta Generata" è "Contesto non sufficiente.", assegna 1 a tutti i punteggi.
    """

    user_prompt = f"""
    Valuta la seguente risposta fornendo un JSON con i 6 campi richiesti:
    
    [Domanda Originale]:
    {domanda}
    
    [Risposta Ideale (riferimento)]:
    {risposta_ideale}
    
    [Risposta Generata (da valutare)]:
    {risposta_generata}
    """
    
    default_error_response = {
        "punteggio_chiarezza": 0, "punteggio_rilevanza": 0, "punteggio_correttezza": 0,
        "punteggio_completezza": 0, "punteggio_coerenza": 0, "giustificazione_generale": "Errore durante la valutazione"
    }

    try:
        response = judge_client.chat.completions.create(
            model=JUDGE_MODEL_DEPLOYMENT,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0,
            response_format={"type": "json_object"}
        )
        
        result_json = json.loads(response.choices[0].message.content)
        for key in default_error_response.keys():
            if key not in result_json:
                result_json[key] = 0 if 'punteggio' in key else "Mancante"
        return result_json

    except Exception as e:
        print(f"    -> ERRORE durante la chiamata al modello giudice: {e}")
        default_error_response["giustificazione_generale"] = f"Errore: {e}"
        return default_error_response

def formatta_e_salva_report(df: pd.DataFrame, file_path: str):
    """
    Salva un DataFrame in un file Excel con la formattazione richiesta dall'utente.
    """
    print("\nSalvataggio del report formattato su file Excel...")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Report Valutazione Giudice"

        # Popola il foglio Excel con i dati del DataFrame, inclusa l'intestazione
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # --- Stili di Formattazione (dallo snippet dell'utente) ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        # Stile per le righe con punteggio 0 (rosso chiaro)
        score_zero_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Applica stile all'intestazione
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Trova l'indice della colonna "Punteggio Similarità"
        punteggio_col_idx = -1
        for idx, col_name in enumerate(df.columns, 1):
            if col_name == "Punteggio Similarità":
                punteggio_col_idx = idx
                break

        # Applica stile alle righe di dati
        for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
            # Applica stile di base a tutte le celle della riga
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top") # Allineamento in alto come da snippet
                cell.border = thin_border
            
            # --- LOGICA DI EVIDENZIAZIONE IN ROSSO ---
            if punteggio_col_idx != -1:
                punteggio_cell = ws.cell(row=row_index, column=punteggio_col_idx)
                try:
                    if float(punteggio_cell.value) == 0.0:
                        for cell in row:
                            cell.fill = score_zero_fill
                except (ValueError, TypeError):
                    pass

        # Imposta larghezza delle colonne (adattate per le nuove colonne)
        column_widths = {
            'A': 20, 'B': 50, 'C': 50, 'D': 60, 'E': 60, 'F': 20, # Colonne originali
            'G': 20, 'H': 20, 'I': 20, 'J': 20, 'K': 20, 'L': 60  # Colonne del giudice
        }
        col_letters = [chr(ord('A') + i) for i in range(len(df.columns))]
        for i, col_letter in enumerate(col_letters):
            if col_letter in column_widths:
                 ws.column_dimensions[col_letter].width = column_widths[col_letter]

        wb.save(file_path)
        print(f"✅ Report salvato con successo come '{file_path}'")

    except Exception as e:
        print(f"ERRORE durante il salvataggio del report Excel: {e}")


# --- Flusso Principale ---
if __name__ == "__main__":
    print(f"--- INIZIO VALUTAZIONE DETTAGLIATA CON GIUDICE LLM ---")
    print(f"Lettura del report di input da: '{FILE_REPORT_DA_VALUTARE}'")

    try:
        df_full = pd.read_excel(FILE_REPORT_DA_VALUTARE)
        df = df_full.head(26).copy()
        print(f"File letto con successo. Elaborazione delle prime {len(df)} righe.")
    except FileNotFoundError:
        print(f"ERRORE: File '{FILE_REPORT_DA_VALUTARE}' non trovato.")
        sys.exit(1)
    except Exception as e:
        print(f"ERRORE durante la lettura del file Excel: {e}")
        sys.exit(1)

    nuove_colonne = [
        'Punteggio Chiarezza', 'Punteggio Rilevanza', 'Punteggio Correttezza',
        'Punteggio Completezza', 'Punteggio Coerenza', 'Giustificazione Giudice'
    ]
    for col in nuove_colonne:
        df[col] = ''

    for index, row in df.iterrows():
        print(f"\nValutazione riga {index + 1}/{len(df)}...")
        
        domanda = row['Domanda Originale']
        risposta_ideale = row['Risposta Ideale']
        risposta_generata = row['Risposta Generata']
        
        if pd.isna(risposta_generata):
            print("  -> Risposta generata vuota. Salto.")
            df.at[index, 'Giustificazione Giudice'] = "Risposta mancante nel report"
            continue

        giudizio = ottieni_giudizio_dettagliato_llm(domanda, risposta_ideale, risposta_generata)
        
        df.at[index, 'Punteggio Chiarezza'] = giudizio.get('punteggio_chiarezza')
        df.at[index, 'Punteggio Rilevanza'] = giudizio.get('punteggio_rilevanza')
        df.at[index, 'Punteggio Correttezza'] = giudizio.get('punteggio_correttezza')
        df.at[index, 'Punteggio Completezza'] = giudizio.get('punteggio_completezza')
        df.at[index, 'Punteggio Coerenza'] = giudizio.get('punteggio_coerenza')
        df.at[index, 'Giustificazione Giudice'] = giudizio.get('giustificazione_generale')
        print(f"  -> Risultato: {giudizio}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = nome_file_report.split('.xlsx')[0]
    output_filename = f"{base_name}_con_giudizio_formattato_{timestamp}.xlsx"
    percorso_output = os.path.join(script_directory, output_filename)
    
    # Chiama la nuova funzione di formattazione
    formatta_e_salva_report(df, percorso_output)